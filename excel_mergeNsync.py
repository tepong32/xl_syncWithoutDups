from openpyxl import load_workbook
from datetime import datetime
import shutil
import re
import hashlib
from decimal import Decimal, InvalidOperation


MASTER_FILE = "official.xlsx"
INCOMING_FILE = "daily_extract.xlsx"
SHEET = "Sheet1"

UPDATE_COLS = ["A", "B", "C", "D", "E", "F", "G"]

# Backup
backup = f"official_backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
shutil.copy(MASTER_FILE, backup)

wb_m = load_workbook(MASTER_FILE)
wb_i = load_workbook(INCOMING_FILE)

ws_m = wb_m[SHEET]
ws_i = wb_i[SHEET]



### logging
LOG_SHEET = "MERGE_LOG"

if LOG_SHEET not in wb_m.sheetnames:
    ws_log = wb_m.create_sheet(LOG_SHEET)
    ws_log.append([
        "Timestamp",
        "Action",
        "Key",
        "CheckNo",
        "RefNo",
        "Reason",
        "IncomingRow"
    ])
else:
    ws_log = wb_m[LOG_SHEET]

def log(action, key="", check_no="", ref_no="", reason="", row=""):
    ws_log.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        action,
        key,
        check_no,
        ref_no,
        reason,
        row
    ])


### sanitation
def clean_text(value):
    if not value:
        return ""
    text = str(value)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().upper()

def clean_number(value):
    try:
        return str(Decimal(str(value)).quantize(Decimal("0.01")))
    except (InvalidOperation, TypeError):
        return "0.00"

def clean_date(value):
    if not value:
        return ""
    try:
        return value.strftime("%Y-%m-%d")
    except AttributeError:
        return str(value).strip()

def sha256_signature(text):
    return hashlib.sha256(text.encode("utf-8")).hexdigest()

def resolve_key(ws, row):
    check_no = ws[f"B{row}"].value
    ref_no   = ws[f"C{row}"].value

    # Strong key: Check + Ref
    if check_no and ref_no:
        return f"CHK:{str(check_no).strip()}|REF:{str(ref_no).strip()}"

    # DV-only fallback with content signature
    if ref_no:
        date  = clean_date(ws[f"A{row}"].value)
        desc  = clean_text(ws[f"E{row}"].value)
        gross = clean_number(ws[f"F{row}"].value)
        net   = clean_number(ws[f"G{row}"].value)

        signature_base = f"{date}|{desc}|{gross}|{net}"
        sig = sha256_signature(signature_base)

        return f"DV:{str(ref_no).strip()}|SIG:{sig}"

    return None


# Build master index
master_index = {}
dv_index = {}

for r in range(2, ws_m.max_row + 1):
    key = resolve_key(ws_m, r)
    if not key:
        continue

    master_index[key] = r

    if key.startswith("DV:"):
        dv_base = key.split("|SIG:")[0]
        dv_index.setdefault(dv_base, []).append(key)

incoming_seen = set()

added = updated = skipped = dup_incoming = dv_conflict = 0

for r in range(2, ws_i.max_row + 1):
    key = resolve_key(ws_i, r)

    check_no = ws_i[f"B{r}"].value
    ref_no   = ws_i[f"C{r}"].value

    if not key:
        skipped += 1
        log(
            "SKIPPED_NO_KEY",
            "",
            check_no,
            ref_no,
            "Missing CheckNo and RefNo",
            r
        )
        continue

    if key in incoming_seen:
        dup_incoming += 1
        log(
            "DUP_INCOMING",
            key,
            check_no,
            ref_no,
            "Duplicate inside incoming file",
            r
        )
        continue
    incoming_seen.add(key)

    # DV conflict detection
    if key.startswith("DV:"):
        dv_base = key.split("|SIG:")[0]
        if dv_base in dv_index and key not in dv_index[dv_base]:
            dv_conflict += 1
            log(
                "DV_CONFLICT",
                key,
                check_no,
                ref_no,
                "Same DV number with different content signature",
                r
            )
            continue

    if key in master_index:
        target = master_index[key]
        for col in UPDATE_COLS:
            ws_m[f"{col}{target}"].value = ws_i[f"{col}{r}"].value

        updated += 1
        log(
            "UPDATED",
            key,
            check_no,
            ref_no,
            "Matched existing record",
            r
        )
    else:
        ws_m.append([ws_i[f"{c}{r}"].value for c in UPDATE_COLS] + [""])
        master_index[key] = ws_m.max_row

        if key.startswith("DV:"):
            dv_index.setdefault(key.split("|SIG:")[0], []).append(key)

        added += 1
        log(
            "ADDED",
            key,
            check_no,
            ref_no,
            "New record appended",
            r
        )

wb_m.save(MASTER_FILE)

print("Merge completed.")
print(f"Added: {added}")
print(f"Updated: {updated}")
print(f"Skipped (no key): {skipped}")
print(f"Skipped (duplicate in incoming): {dup_incoming}")
print(f"DV conflicts: {dv_conflict}")
